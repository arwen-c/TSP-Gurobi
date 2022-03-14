# Import de modules
import numpy as np
import pandas as pd
import math
import openpyxl
from guppy import hpy


# Fonction récupération des données issues des excel


def extractionData(path):
    """Permet l'extraction des données depuis un fichier excel.
    path est une chaîne de caratères correspondant à l'emplacement relatif du fichier.
    Renvoie les différents dictionnaires permettant un traitement des données avec python."""
    xls = pd.ExcelFile(path)
    df1 = pd.read_excel(xls, 'Employees')
    df2 = pd.read_excel(xls, 'Tasks')

    # Création des dictionnaires de données
    EmployeesDico = df1.to_dict('records')
    TasksDico = df2.to_dict('records')

    return EmployeesDico, TasksDico


<<<<<<< HEAD:Phase_1/firstdoc.py
# calcule la distance entre deux points dont on connait les coordonnées GPS


# Gestion de la localisation -- modif de la fonction distance
=======
# Fonctions utiles pour créer les matrices de données utiles

>>>>>>> main:Phase_1/usefulFunctions1.py
def distance(id1, id2, TasksDico):
    """Calcule la distance entre deux points dont on connait les coordonnées GPS.
    Entrée : les taskid correspondantes et le dictionnaire de données.
    Sortie : distance en km."""
    foundId1, foundId2 = False, False
    index = 0
<<<<<<< HEAD:Phase_1/firstdoc.py
    while not (foundid1 and foundid2) and index<len(TasksDico):
=======
    while not (foundId1 and foundId2):
>>>>>>> main:Phase_1/usefulFunctions1.py
        if TasksDico[index]['TaskId'] == id1:
            foundId1 = True
            long1 = TasksDico[index]['Longitude']
            lat1 = TasksDico[index]['Latitude']
        if TasksDico[index]['TaskId'] == id2:
            foundId2 = True
            long2 = TasksDico[index]['Longitude']
            lat2 = TasksDico[index]['Latitude']
        index += 1
<<<<<<< HEAD:Phase_1/firstdoc.py
    if long1 is None or long2 is None:
        distance = 0
    else :
        delta_long = long2-long1  # calcule de la différence de longitude
        delta_latt = lat2-lat1
        distance = (1.852*60*math.sqrt(delta_long**2+delta_latt**2))
=======
    deltaLong = long2-long1  # calcule de la différence de longitude
    deltaLatt = lat2-lat1
    distance = (1.852*60*math.sqrt(deltaLong**2+deltaLatt**2))
>>>>>>> main:Phase_1/usefulFunctions1.py
    return distance

# def distance(id1, id2, TasksDico):
#     '''entrée : les taskid correspondantes, le dictionnaire de données, sortie : distance en km'''
#     foundid1, foundid2 = False, False
#     index = 0
#     while not (foundid1 and foundid2):
#         if TasksDico[index]['TaskId'] == id1:
#             foundid1 = True
#             long1 = TasksDico[index]['Longitude']
#             lat1 = TasksDico[index]['Latitude']
#         if TasksDico[index]['TaskId'] == id2:
#             foundid2 = True
#             long2 = TasksDico[index]['Longitude']
#             lat2 = TasksDico[index]['Latitude']
#         index += 1
#     delta_long = long2-long1  # calcule de la différence de longitude
#     delta_latt = lat2-lat1
#     distance = (1.852*60*math.sqrt(delta_long**2+delta_latt**2))
#     return distance


def competenceOK(EmployeeName, TaskId, TasksDico, EmployeesDico):
    """EmployeeName est une chaîne de caractères correspondant au nom d'un employé.
    TaskId est l'identifiant d'une tâche.
    TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Retourne 1 si l'employé a la bonne compétence et un niveau suffisant pour effectuer la tâche, 0 sinon."""
    taskSkill = next(item['Skill']
                     for item in TasksDico if item['TaskId'] == TaskId)
    employeeSkill = next(
        item['Skill'] for item in EmployeesDico if item['EmployeeName'] == EmployeeName)
    taskLevel = next(item['Level']
                     for item in TasksDico if item['TaskId'] == TaskId)
    employeeLevel = next(
        item['Level'] for item in EmployeesDico if item['EmployeeName'] == EmployeeName)
    if taskLevel <= employeeLevel and taskSkill == employeeSkill:
        return 1
    else:
        return 0


def matrice_distance(dic_taches):
    matrice_des_distances = []
    for tache_1 in dic_taches:
        task_id_1 = tache_1["TaskId"]
        ligne = []
        for tache_2 in dic_taches:
            task_id_2 = tache_2["TaskId"]
            ligne.append(distance(task_id_1, task_id_2, dic_taches))
        matrice_des_distances.append(ligne)
    return matrice_des_distances


def matriceCompetences(EmployeesDico, TasksDico):
    '''Crée une matrice C_ni qui comporte un 1 si l'employé n peut faire la tache i'''
    nombre_employes = len(EmployeesDico)
    nombre_taches = len(TasksDico)
    C = np.zeros((nombre_employes, nombre_taches))
    for n in range(nombre_employes):
        for i in range(nombre_taches):
            C[n, i] = competenceOK(
                EmployeesDico[n]['EmployeeName'], TasksDico[i]['TaskId'], TasksDico, EmployeesDico)
    return C


def vecteurOuvertures(TasksDico):
    '''Crée un vecteur avec les heures d'ouvertures des tâches'''
    nombre_taches = len(TasksDico)
    O = []
    for i in range(nombre_taches):
        heure = TasksDico[i]['OpeningTime']
        res = heure.split(':')
        h = int(res[0])  # l'heure
        m = int(res[1][:1])  # les minutes
        if res[1][2:] == 'pm':
            if res[0]!=12:
                h += 12  # modifications pour l'aprem
        else :
            if res[0]==12 : h=0
        O.append(h*60+m)
    return O


def vecteurFermetures(TasksDico):
    '''Crée un vecteur avec les heures de fermeture des tâches'''
    nombre_taches = len(TasksDico)
    F = []
    for i in range(nombre_taches):
        heure = TasksDico[i]['ClosingTime']
        res = heure.split(':')
        h = int(res[0])  # l'heure
        m = int(res[1][:1])  # les minutes
        if res[1][2:] == 'pm':
            if res[0]!=12:
                h += 12  # modifications pour l'aprem
        else :
            if res[0]==12 : h=0
        F.append(h*60+m)
    return F


def vecteurDurees(TasksDico):
    """Pas besoin d'argument, on utilise les variables globales.
    Renvoie le vecteur des durées de chaque tâche."""
    Vecteur_duree = []
    for row in TasksDico:
        Vecteur_duree.append(row['TaskDuration'])
    return Vecteur_duree


# Création du fichier solution


def nom_fichier_resolution(nom_fichier, n_methode):
    """nom_fichier est un string qui correspond au nom du fichier excel.
    n_methode est le numéro de la méthode.
    Renvoie le nom du fichier txt."""
    L = nom_fichier.split('.')
    n_methode = str(n_methode)
    # return 'Solution' + L[0][8:] + 'ByV' + n_methode + '.txt' # ancienne version sans arborescence pour les fichiers excels
    # nouvelle version avec arborescence pour les fichiers excels
    return 'Phase_1/Solutions/Solution' + L[0][28:] + 'ByV' + n_methode + '.txt'


def solution_fichier_txt(X, h, EmployeesDico):
    """X et h sont des tableaux numpy.
    X est de dimension 3 et h de dimension 1.
    Renvoie la liste des lignes sous la forme souhaitée pour le fichier .txt."""
    premiere_ligne = 'taskId;performed;employeeName;startTime;'
    liste_des_lignes = [premiere_ligne]
    n, x, y = X.shape
    employeeName = ''
    for i in range(x - 2*n):
        j = 0
        tache_i_ajoutee = False
        while j < y and not(tache_i_ajoutee):
            numero_employe = 0
            while numero_employe < n and not(tache_i_ajoutee):
                if X[numero_employe, i, j] == 1:
                    employeeName = EmployeesDico[numero_employe]['EmployeeName']
                    liste_des_lignes.append(
                        'T' + str(i+1) + ';' + '1' + ';' + str(employeeName) + ';' + str(round(h[i])) + ';')
                    tache_i_ajoutee = True
                numero_employe = numero_employe + 1
            j = j + 1
        if not(tache_i_ajoutee):
            liste_des_lignes.append(
                'T' + str(i+1) + ';' + '0' + ';' + ';' + ';')
    return liste_des_lignes


def creation_fichier(nom_fichier, n_methode, X, h, EmployeesDico):
    """nom_fichier est le nom du fichier utilisée pour créer les variables globales.
    n_methode est le numéro de la méthode utilisée.
    X est un tableau en 3 dimensions où chaque coefficient permet de savoir si l'employé n est allé de la tâche i à la tâche j.
    Ne renvoie rien mais crée ou modifie le fichier .txt."""
    file = open(nom_fichier_resolution(nom_fichier, n_methode), "w")
    file.write("\n".join(solution_fichier_txt(X, h, EmployeesDico)))
    file.close()
    return None

def performances1 (tpsExec, tailleEntree, tailleMemoire):
    #Ecriture des critères de performance dans un excel
    my_path = ".\performance1.xlsx"
    my_wb = openpyxl.load_workbook(my_path)
    my_sheet = my_wb.active
    #on cherche à partir de quelle ligne écrire (écriture à la suite)
    i=0
    cell=my_sheet.cell(row = i, column = 1)
    while cell.value !=None:
        i+=1
    #on ajoute les valeurs de performance obtenue du code
    cell.value =tpsExec #temps d'execution en première colonne
    cell=my_sheet.cell(row = i, column = 2)
    cell.value =tailleEntree #taille des instances d'entrée en deuxième colonne
    cell=my_sheet.cell(row = i, column = 3)
    cell.value =tailleMemoire #taille de la mémoire occupée par le programme en troisième colonne
    cell=my_sheet.cell(row = i, column = 6)
    cell.value =phase #type de phase ayant conduit à la valeur donnée (phase 1 ou 2)
    #on calcule àpartir de ces valeurs de nouveaux indicateurs
    cell=my_sheet.cell(row = i, column = 4)
    cell.value =tpsExec/tailleEntree
    cell=my_sheet.cell(row = i, column = 5)
    cell.value =tailleMemoire/tailleEntree
    #on enregistre les données au sein de l'excel
    my_wb.save(".\performance1.xlsx")