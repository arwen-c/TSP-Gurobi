import math
import numpy as np
import pandas as pd
import math
import openpyxl


def fonctionRestriction(capaciteEmploye, X):
    """capaciteEmploye = matrice des compétences
    X = matrice d'affectation des tâches, l'employé n va de la tâche i à la tâche j
    Cette fonction permet de restreindre la recherche de taches que l'employé va faire en vérifiant la contrainte de compétences et si la tâche a déjà été faite.
    Renvoie une liste d'identifiant de tâches.
    """
    listeTaches = []
    # si on garde 3 dim : nbreEmploye, nbreTache, _ = X.shape
    nbreEmploye, nbreTache = X.shape
    for k in range(0, nbreTache):
        if capaciteEmploye[k] == 1:  # car le tableau des compétences commencent à 0
            tacheNonFaite = True
            n = 0
            while tacheNonFaite and n < nbreEmploye:
                # si X a 3 dim prendre le code ci-dessous
                # i = 0
                # while tacheNonFaite and i < nbreTache:
                #     if X[n, i, k] == 1:  # or X[n, k, i] == 1: normalement pas besoin de regarder dans les 2 sens
                #         tacheNonFaite = False
                #     i += 1
                # n += 1
                if X[n, k] == 1:  # même chose que pour compétence
                    tacheNonFaite = False
                n += 1
            if tacheNonFaite:
                listeTaches.append(k)
    # peut-être renvoyer une liste de dictionnaires pour travailler avec une instance plus petite dans la fonction glouton et donc gagner en complexité (mais peut-être inutile)
    return listeTaches


def triOpti(tachesFaisables, localisationCourante, matDistance, duree):
    '''
    Fonction ayant pour objectif de trier les numéros des taches en fonction de leur optimalité vis-à-vis du critère d'optimisation local choisi
    ENTREES
    - tachesFaisables : liste des indices des taches faisables
    - localisationCourante : indice de la dernière tache effectuée par l'employé (permet de déterminer sa position comme étant celle de la dernière tache effectuée)
    - distance : matrice contenant les distances entre la tache d'indice colonne et la tache d'indice ligne
    SORTIES
    - tableau trié des taches de la plus avantageuse à la moins avantageuse
    selon quel critère optimiser ?
    Choix d'optimiser en fonction de l'argent gagné = cout horaire du dépanage*duree-cout transport*distance
    '''
    # création d'une liste contenant les valeurs de cout associées à la tache
    L = len(tachesFaisables)
    cout = L*[0]
    for i in range(L):
        cout[i] = (2/3)*duree[i] + (12/0.833-(0.575+0.12)) * \
            matDistance[tachesFaisables[i]][localisationCourante]
    tachesFaisables = list(tachesFaisables.copy())
    coutTrie = np.sort(cout.copy())
    tacheTrie = L*[0]
    for i in range(L):
        j = 0
        trouve = False
        while not(trouve) and j < len(tachesFaisables):
            if cout[j] == coutTrie[i]:
                tacheTrie[i] = tachesFaisables.pop(j)
                cout.pop(j)
                trouve = True
            j += 1
    # souci d'une tache qui n'est plus dans tacheTrie si deux taches ont le même cout. On peut faire un point pop, mais perte en lisibilité de programme non ?
    return tacheTrie


def tachesRealisables(tachesOpti, duree, debut, fin, finJourneeEmploye, indispoDicoEmployeN, t, pauseFaite, localisationCourante, matDistance, tachesDico):
    '''
    tachesOpti = liste des tâches triées selon l'optimisation de l'objectif ;
    duree = vecteur durée des tâches ;
    debut = vecteur avec les ouvertures des tâches après chaque indisponibilité ;
    fin = vecteur avec les ouvertures des tâches avant chaque indisponibilité ;
    finJourneeEmploye = heure de fin de la journée de l'employé sélectionné ;
    indispoDico = dictionnaire qui regroupe les informations relatives aux indisponibilités de l'emplyé sélectionné.

    Cette fonction vérifie qu'au moins une tâche est réalisable par l'employé avant la fin de la journée, qui retourne :
    - raison valant 'indisponibilité', si une indisponibilité bloque la réalisation d'une de ces tâches, 'fin de journée' si aucun tâche n'est faisable avant la fin de journée, 'déjeuner' si le déjeuner bloque, none sinon
    - tache : le numéro de la tache optimale faisable, si une tache au moins est faisable, none sinon
    '''
    # comment prendre en compte les ouvertures et les fermetures ?
    # heureButtoire = min(heure de fin, heure d'indispo - si existe-, Hfin de pause dej-temps de dej si pause non faite)

    raison = ''
    tache = None
    tacheOptiFaisableTrouvee = False
    m = len(tachesOpti)
    k = 0
    while not(tacheOptiFaisableTrouvee) and k < m:
        if finJourneeEmploye > t + matDistance[localisationCourante][tachesOpti[k]]/0.833 + duree[tachesOpti[k]] + matDistance[tachesOpti[k]][k]/0.833:
            # nombre de créneaux de disponibilité pour une tâche donnée
            nbreCreneauxDebutK = len(debut[k])
            # on vérifie que des créneaux d'ouverture des tâches sont suffisamment grand
            creneauConvenable = False
            c = 0
            t2 = 0
            while not(creneauConvenable) and c < nbreCreneauxDebutK:
                if t + matDistance[localisationCourante][tachesOpti[k]]/0.833 + duree[k] < fin[k][c]:
                    # soit le créneau est déjà ouvert, soit on peut attendre son ouverture ()
                    if t + matDistance[localisationCourante][tachesOpti[k]]/0.833 > debut[k][c]:
                        creneauConvenable = True
                    elif fin[k][c]-duree[k] > 0:
                        creneauConvenable = True
                        t2 = debut[k][c]-t
                c += 1

                # # + 10: # le +10 peermet de ne pas rater une tache optimale à quelques minutes près 10 en l'occurence ici
                # if t + matDistance[localisationCourante][tachesOpti[k]]/0.833 > debut[k][c]:
                #     # + 10 en fait pb il faudrait savoir si on a bien rajouté ces 10 min boucle while  # on regarde la fin de créneau correspondante + il faut faire attention au décalage qu'on a pu créer précédemment avec le + 10
                #     if t + matDistance[localisationCourante][tachesOpti[k]]/0.833 + duree[k] < fin[k][c]:
                #         creneauConvenable = True

            # on vérifie que notre employé est disponible à l'un de ces créneaux
            pasIndispo = False
            if creneauConvenable and indispoDicoEmployeN != {}:
                distancePourIndispo = distanceGPS(tachesDico[localisationCourante]['Latitude'], tachesDico[localisationCourante]
                                                  ['Longitude'], indispoDicoEmployeN['Latitude'], indispoDicoEmployeN['Longitude'])
                pasIndispo = recuperationHeure(
                    indispoDicoEmployeN['Start']) > t + matDistance[localisationCourante][tachesOpti[k]]/0.833 + duree[k] + distancePourIndispo/0.833
                if not(pasIndispo):  # si on a bien une indisponibilité
                    raison = 'indisponibilité'
            if creneauConvenable and pasIndispo:
                if pauseFaite:
                    tache = int(tachesOpti[k])
                    tacheOptiFaisableTrouvee = True
                else:
                    if t + t2 + matDistance[localisationCourante][tachesOpti[k]]/0.833 + duree[tachesOpti[k]] < 780:
                        tache = int(tachesOpti[k])
                        tacheOptiFaisableTrouvee = True
                    else:
                        raison = 'déjeuner'
        k += 1

    if raison == '' and not tacheOptiFaisableTrouvee:
        raison = 'fin de journée'

    return raison, tache


def extractionData(path):
    """Permet l'extraction des données depuis un fichier excel.
    path est une chaîne de caratères correspondant à l'emplacement relatif du fichier.
    Renvoie les différents dictionnaires permettant un traitement des données avec python."""
    xls = pd.ExcelFile(path)
    df1 = pd.read_excel(xls, 'Employees')
    df2 = pd.read_excel(xls, 'Employees Unavailabilities')
    df3 = pd.read_excel(xls, 'Tasks')
    df4 = pd.read_excel(xls, 'Tasks Unavailabilities')

    # Création des dictionnaires de données
    EmployeesDico = df1.to_dict('records')
    EmployeesUnavailDico = df2.to_dict('records')
    TasksDico = df3.to_dict('records')
    TasksUnavailDico = df4.to_dict('records')

    return EmployeesDico, EmployeesUnavailDico, TasksDico, TasksUnavailDico


# Fonctions utiles pour créer les matrices de données utiles
def deg2rad(dd):
    """Convertit un angle "degrés décimaux" en "radians"
    """
    return dd/180*math.pi


def distanceGPS(latA, longA, latB, longB):
    """Retourne la distance en mètres entre les 2 points A et B connus grâce à
       leurs coordonnées GPS (en radians).
    """
    # Rayon de la terre en kilomètres (sphère IAG-GRS80)
    RT = 6378.137
    # angle en radians entre les 2 points
    x = math.sin(latA)*math.sin(latB) + math.cos(latA) * \
        math.cos(latB)*math.cos(abs(longB-longA))
    if abs(x-1) <= 0.000000000001:
        x = 1
    elif abs(x+1) <= 0.000000000001:
        x = -1
    S = math.acos(x)
    # distance entre les 2 points, comptée sur un arc de grand cercle - en km
    return S*RT


def distance(id1, id2, TasksDico):
    """Calcule la distance entre deux points dont on connait les coordonnées GPS.
    Entrée : les taskid correspondantes et le dictionnaire de données.
    Sortie : distance en km."""
    foundId1, foundId2 = False, False
    index = 0

    while not (foundId1 and foundId2):
        if TasksDico[index]['TaskId'] == id1:
            foundId1 = True
            long1 = deg2rad(TasksDico[index]['Longitude'])
            lat1 = deg2rad(TasksDico[index]['Latitude'])
        if TasksDico[index]['TaskId'] == id2:
            foundId2 = True
            long2 = deg2rad(TasksDico[index]['Longitude'])
            lat2 = deg2rad(TasksDico[index]['Latitude'])
        index += 1

    # distance d'arc entre deux points
    distance = distanceGPS(lat1, long1, lat2, long2)/1000
    return distance


def competenceOK(EmployeeName, TaskId, TasksDico, EmployeesDico):
    """EmployeeName est une chaîne de caractères correspondant au nom d'un employé.
    TaskId est l'identifiant d'une tâche.
    TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Retourne 1 si l'employé a la bonne compétence et un niveau suffisant pour effectuer la tâche, 0 sinon."""
    taskSkill = next(item['Skill']
                     for item in TasksDico if item['TaskId'] == TaskId)
    employeeSkill = next(
        item['Skill'] for item in EmployeesDico if item['EmployeeName'] == EmployeeName)
    taskLevel = next(item['Level']
                     for item in TasksDico if item['TaskId'] == TaskId)
    employeeLevel = next(
        item['Level'] for item in EmployeesDico if item['EmployeeName'] == EmployeeName)
    if taskLevel <= employeeLevel and taskSkill == employeeSkill:
        return 1
    elif taskSkill == None:
        return 1
    else:
        return 0


# Fonctions création des matrices nécessaires au calcul de la fonction coût et à la vérification des contraintes

def matriceDistance(TasksDico):
    """TasksDico est le dictionnaire contenant les informations sur les tâches.
    Renvoie la matrice des distances entre les différentes tâches."""
    matriceDesDistances = []
    for tache1 in TasksDico:
        taskId1 = tache1["TaskId"]
        ligne = []
        for tache2 in TasksDico:
            taskId2 = tache2["TaskId"]
            ligne.append(distance(taskId1, taskId2, TasksDico))
        matriceDesDistances.append(ligne)
    return matriceDesDistances


def matriceCompetences(EmployeesDico, TasksDico):
    """TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Renvoie une matrice C qui comporte un 1 si l'employé n peut faire la tache i et un 0 sinon."""
    nombreEmployes = len(EmployeesDico)
    nombreTaches = len(TasksDico)
    C = np.zeros((nombreEmployes, nombreTaches))
    for n in range(nombreEmployes):
        for i in range(nombreTaches):
            C[n, i] = competenceOK(
                EmployeesDico[n]['EmployeeName'], TasksDico[i]['TaskId'], TasksDico, EmployeesDico)
    return C


# Fonctions récupération des contraintes temporelles des tâches

def recuperationHeure(heure):
    """heure est une chaîne de caractères de la forme 11:15pm.
    Renvoie un entier correspondant au nombre de minutes écoulées depuis minuit."""
    res = heure.split(':')
    h = int(res[0])  # l'heure
    m = int(res[1][:2])  # les minutes
    if res[1][2:] == 'pm':
        if res[0] != 12:
            h += 12  # modifications pour l'aprem
    else:
        if res[0] == 12:
            h = 0
    return h*60+m


def vecteurOuvertures(TasksDico, TasksUnavailDico):
    """TasksDico et TasksUnavailDico sont les dictionnaires contenant les informations respectives sur les tâches et leurs indisponibilités.
    Renvoie une liste de liste avec les heures d'ouvertures de chaque tâche après chaque indisponibilité."""
    nombreTaches = len(TasksDico)
    nombreIndisponibilite = len(TasksUnavailDico)
    O = []
    for i in range(nombreTaches):
        horairesOvertures = []
        heure = TasksDico[i]['OpeningTime']
        horairesOvertures.append(recuperationHeure(heure))
        for k in range(nombreIndisponibilite):
            # on vérifie que l'indisponibilité correspond à la tâche i
            if TasksDico[i]['TaskId'] == TasksUnavailDico[k]['TaskId']:
                heure = TasksUnavailDico[k]['End']
                # ajout de l'heure de réouverture après l'indisponibilité k de la tâche i
                horairesOvertures.append(recuperationHeure(heure))
        O.append(horairesOvertures)
    return O


def vecteurFermetures(TasksDico, TasksUnavailDico):
    """TasksDico et TasksUnavailDico sont les dictionnaires contenant les informations respectives sur les tâches et leurs indisponibilités.
    Renvoie une liste de liste avec les heures de fermetures de chaque tâche avant chaque indisponibilité."""
    nombreTaches = len(TasksDico)
    nombreIndisponibilites = len(TasksUnavailDico)
    F = []
    for i in range(nombreTaches):
        horairesFermetures = []
        heure = TasksDico[i]['ClosingTime']
        horairesFermetures.append(recuperationHeure(heure))
        for k in range(nombreIndisponibilites):
            # on vérifie que l'indisponibilité correspond à la tâche i
            if TasksDico[i]['TaskId'] == TasksUnavailDico[k]['TaskId']:
                heure = TasksUnavailDico[k]['Start']
                # ajout de l'heure de réouverture après l'indisponibilité k de la tâche i
                horairesFermetures.append(recuperationHeure(heure))
        F.append(horairesFermetures)
    return F


def vecteurDurees(TasksDico):
    """TasksDico est le dictionnaire contenant les informations sur les tâches.
    Renvoie le vecteur des durées de chaque tâche."""
    Vecteur = []
    for row in TasksDico:
        Vecteur.append(row['TaskDuration'])
    return Vecteur


# Creation de taches fictives pour utiliser notre modèle

def ajoutTachesFictives(TasksDico, EmployeesDico, EmployeesUnavailDico):

    # (rechercher ajout_domicile dans tous les docs pour modifier par ajout_taches_fictives)
    ### AJOUTS DOMICILES ###
    """Modification des données pour insérer des tâches factices de départ et de retour au dépot ou domicile."""
    TasksEnhanced = TasksDico.copy()
    for row in EmployeesDico:
        # ajout d'une tâche au départ du domicile
        TasksEnhanced.append({'TaskId': 'Depart' + row['EmployeeName'], 'Latitude': row['Latitude'],    'Longitude': row['Longitude'],
                              'TaskDuration': 0, 'Skill': row['Skill'], 'Level': 0, 'OpeningTime': row['WorkingStartTime'], 'ClosingTime': row['WorkingEndTime']})
    return TasksEnhanced


# Fonctions création du fichier solution

def nomFichierResolution(nomFichier, nMethode):
    """nomFichier est un string qui correspond au nom du fichier excel.
    n_methode est le numéro de la méthode.
    Renvoie le nom du fichier txt."""
    L = nomFichier.split('.')
    return 'Phase3/Solutions/Solution' + L[0][27:] + 'ByV' + str(nMethode) + '.txt'


def lignesSolution(X, h, L, TasksDico, EmployeesDico):
    """X et h sont des tableaux numpy.
    X est de dimension 3 et h de dimension 1.
    Renvoie la liste des lignes sous la forme souhaitée pour le fichier .txt."""
    # Duree = vecteurDurees(TasksDico)
    premiereLigne = 'taskId;performed;employeeName;startTime;'
    listeDesLignes = [premiereLigne]
    n, _ = X.shape
    employeeName = ''
    nombreTaches = len(TasksDico)
    for i in range(nombreTaches):
        tacheiAjoutee = False
        numeroEmploye = 0
        while numeroEmploye < n and not(tacheiAjoutee):
            if X[numeroEmploye, i] == 1:
                employeeName = EmployeesDico[numeroEmploye]['EmployeeName']
                listeDesLignes.append(
                    'T' + str(i+1) + ';' + '1' + ';' + str(employeeName) + ';' + str(round(h[i])) + ';')
                tacheiAjoutee = True
            numeroEmploye = numeroEmploye + 1
        if not(tacheiAjoutee):
            listeDesLignes.append('T' + str(i+1) + ';' + '0' + ';' + ';' + ';')
    listeDesLignes.append(' ')  # saut de ligne
    listeDesLignes.append('employeeName;lunchBreakStartTime;')
    for numeroEmploye in range(n):
        listeDesLignes.append(str(
            EmployeesDico[numeroEmploye]['EmployeeName']) + ';' + str(L[numeroEmploye]) + ';')
    return listeDesLignes


def creationFichier(nomFichier, nMethode, X, h, L, TasksDico, EmployeesDico):
    """nomFichier est le nom du fichier utilisée pour créer les variables globales.
    n_methode est le numéro de la méthode utilisée.
    X est un tableau en 3 dimensions où chaque coefficient permet de savoir si l'employé n est allé de la tâche i à la tâche j.
    Ne renvoie rien mais crée ou modifie le fichier .txt."""
    fichier = open(nomFichierResolution(nomFichier, nMethode), "w")
    fichier.write("\n".join(lignesSolution(
        X, h, L, TasksDico, EmployeesDico)))
    fichier.close()
    return None


def performances2(tpsExec, tailleEntree, tailleMemoire, instance):
    # Ecriture des critères de performance dans un excel
    my_path = "./performance1.xlsx"
    my_wb = openpyxl.load_workbook(my_path)
    my_sheet = my_wb.active
    # on cherche à partir de quelle ligne écrire (écriture à la suite)
    i = 4
    cell = my_sheet.cell(row=i, column=1)
    while cell.value != None:
        i += 1
        cell = my_sheet.cell(row=i, column=1)
    # on ajoute les valeurs de performance obtenue du code
    cell.value = tpsExec  # temps d'execution en première colonne
    cell = my_sheet.cell(row=i, column=2)
    cell.value = tailleEntree  # taille des instances d'entrée en deuxième colonne
    cell = my_sheet.cell(row=i, column=3)
    # taille de la mémoire occupée par le programme en troisième colonne
    cell.value = tailleMemoire
    # on calcule àpartir de ces valeurs de nouveaux indicateurs
    cell = my_sheet.cell(row=i, column=4)
    cell.value = tpsExec/tailleEntree
    cell = my_sheet.cell(row=i, column=5)
    cell.value = tailleMemoire/tailleEntree
    cell = my_sheet.cell(row=i, column=6)
    cell.value = instance
    # on enregistre les données au sein de l'excel
    my_wb.save("./performance2.xlsx")
