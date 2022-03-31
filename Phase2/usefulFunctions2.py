# Import de modules
import numpy as np
import pandas as pd
import math
import openpyxl


# Fonction récupération des données issues des excel

def extractionData(path):
    """Permet l'extraction des données depuis un fichier excel.
    path est une chaîne de caratères correspondant à l'emplacement relatif du fichier.
    Renvoie les différents dictionnaires permettant un traitement des données avec python."""
    xls = pd.ExcelFile(path)
    df1 = pd.read_excel(xls, 'Employees')
    df2 = pd.read_excel(xls, 'Employees Unavailabilities')
    df3 = pd.read_excel(xls, 'Tasks')
    df4 = pd.read_excel(xls, 'Tasks Unavailabilities')

    # Création des dictionnaires de données
    EmployeesDico = df1.to_dict('records')
    EmployeesUnavailDico = df2.to_dict('records')
    TasksDico = df3.to_dict('records')
    TasksUnavailDico = df4.to_dict('records')

    return EmployeesDico, EmployeesUnavailDico, TasksDico, TasksUnavailDico


# Fonctions utiles pour créer les matrices de données utiles

def deg2rad(dd):
    """Convertit un angle "degrés décimaux" en "radians"
    """
    return dd/180*math.pi

def distanceGPS(latA, longA, latB, longB):
    """Retourne la distance en mètres entre les 2 points A et B connus grâce à
       leurs coordonnées GPS (en radians).
    """
    # Rayon de la terre en mètres (sphère IAG-GRS80)
    RT = 6378137
    # angle en radians entre les 2 points
    x = math.sin(latA)*math.sin(latB) + math.cos(latA)*math.cos(latB)*math.cos(abs(longB-longA))        
    if abs(x-1) <= 0.000000000001:
        x = 1
    elif abs(x+1) <= 0.000000000001:
        x = -1
    S = math.acos(x)
    # distance entre les 2 points, comptée sur un arc de grand cercle
    return S*RT
    
def distance(id1, id2, TasksDico):
    """Calcule la distance entre deux points dont on connait les coordonnées GPS.
    Entrée : les taskid correspondantes et le dictionnaire de données.
    Sortie : distance en km."""
    foundId1, foundId2 = False, False
    index = 0
    while not (foundId1 and foundId2):
        if TasksDico[index]['TaskId'] == id1:
            foundId1 = True
            long1 = deg2rad(TasksDico[index]['Longitude'])
            lat1 = deg2rad(TasksDico[index]['Latitude'])
        if TasksDico[index]['TaskId'] == id2:
            foundId2 = True
            long2 = deg2rad(TasksDico[index]['Longitude'])
            lat2 = deg2rad(TasksDico[index]['Latitude'])
        index += 1

    #distance d'arc entre deux points
    distance = round(distanceGPS(lat1,long1,lat2,long2)/1000,1)
    return distance


def competenceOK(EmployeeName, TaskId, TasksDico, EmployeesDico):
    """EmployeeName est une chaîne de caractères correspondant au nom d'un employé.
    TaskId est l'identifiant d'une tâche.
    TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Retourne 1 si l'employé a la bonne compétence et un niveau suffisant pour effectuer la tâche, 0 sinon."""
    taskSkill = next(item['Skill']
                     for item in TasksDico if item['TaskId'] == TaskId)
    employeeSkill = next(
        item['Skill'] for item in EmployeesDico if item['EmployeeName'] == EmployeeName)
    taskLevel = next(item['Level']
                     for item in TasksDico if item['TaskId'] == TaskId)
    employeeLevel = next(
        item['Level'] for item in EmployeesDico if item['EmployeeName'] == EmployeeName)
    if taskLevel <= employeeLevel and taskSkill == employeeSkill:
        return 1
    elif taskSkill == None:
        return 1
    else:
        return 0


# Fonctions création des matrices nécessaires au calcul de la fonction coût et à la vérification des contraintes

def matriceDistance(TasksDico):
    """TasksDico est le dictionnaire contenant les informations sur les tâches.
    Renvoie la matrice des distances entre les différentes tâches."""
    matriceDesDistances = []
    for tache1 in TasksDico:
        taskId1 = tache1["TaskId"]
        ligne = []
        for tache2 in TasksDico:
            taskId2 = tache2["TaskId"]
            ligne.append(distance(taskId1, taskId2, TasksDico))
        matriceDesDistances.append(ligne)
    return matriceDesDistances


def matriceCompetences(EmployeesDico, TasksDico):
    """TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Renvoie une matrice C qui comporte un 1 si l'employé n peut faire la tache i et un 0 sinon."""
    nombreEmployes = len(EmployeesDico)
    nombreTaches = len(TasksDico)
    C = np.zeros((nombreEmployes, nombreTaches))
    for n in range(nombreEmployes):
        for i in range(nombreTaches):
            C[n, i] = competenceOK(
                EmployeesDico[n]['EmployeeName'], TasksDico[i]['TaskId'], TasksDico, EmployeesDico)
    return C


# Fonctions récupération des contraintes temporelles des tâches

def recuperationHeure(heure):
    """heure est une chaîne de caractères de la forme 11:15pm.
    Renvoie un entier correspondant au nombre de minutes écoulées depuis minuit."""
    res = heure.split(':')
    h = int(res[0])  # l'heure
    m = int(res[1][:2])  # les minutes
    if res[1][2:] == 'pm':
        if res[0] != 12:
            h += 12  # modifications pour l'aprem
    else:
        if res[0] == 12:
            h = 0
    return h*60+m


def vecteurOuvertures(TasksDico, TasksUnavailDico):
    """TasksDico et TasksUnavailDico sont les dictionnaires contenant les informations respectives sur les tâches et leurs indisponibilités.
    Renvoie une liste de liste avec les heures d'ouvertures de chaque tâche après chaque indisponibilité."""
    nombreTaches = len(TasksDico)
    nombreIndisponibilite = len(TasksUnavailDico)
    O = []
    for i in range(nombreTaches):
        horairesOvertures = []
        heure = TasksDico[i]['OpeningTime']
        horairesOvertures.append(recuperationHeure(heure))
        for k in range(nombreIndisponibilite):
            # on vérifie que l'indisponibilité correspond à la tâche i
            if TasksDico[i]['TaskId'] == TasksUnavailDico[k]['TaskId']:
                heure = TasksUnavailDico[k]['End']
                # ajout de l'heure de réouverture après l'indisponibilité k de la tâche i
                horairesOvertures.append(recuperationHeure(heure))
        O.append(horairesOvertures)
    return O


def vecteurFermetures(TasksDico, TasksUnavailDico):
    """TasksDico et TasksUnavailDico sont les dictionnaires contenant les informations respectives sur les tâches et leurs indisponibilités.
    Renvoie une liste de liste avec les heures de fermetures de chaque tâche avant chaque indisponibilité."""
    nombreTaches = len(TasksDico)
    nombreIndisponibilites = len(TasksUnavailDico)
    F = []
    for i in range(nombreTaches):
        horairesFermetures = []
        heure = TasksDico[i]['ClosingTime']
        horairesFermetures.append(recuperationHeure(heure))
        for k in range(nombreIndisponibilites):
            # on vérifie que l'indisponibilité correspond à la tâche i
            if TasksDico[i]['TaskId'] == TasksUnavailDico[k]['TaskId']:
                heure = TasksUnavailDico[k]['Start']
                # ajout de l'heure de réouverture après l'indisponibilité k de la tâche i
                horairesFermetures.append(recuperationHeure(heure))
        F.append(horairesFermetures)
    return F


def vecteurDurees(TasksDico):
    """TasksDico est le dictionnaire contenant les informations sur les tâches.
    Renvoie le vecteur des durées de chaque tâche."""
    Vecteur = []
    for row in TasksDico:
        Vecteur.append(row['TaskDuration'])
    return Vecteur


# Fonctions création du fichier solution

def nomFichierResolution(nomFichier, nMethode):
    """nomFichier est un string qui correspond au nom du fichier excel.
    n_methode est le numéro de la méthode.
    Renvoie le nom du fichier txt."""
    L = nomFichier.split('.')
    return 'Phase2/Solutions/Solution' + L[0][27:] + 'ByV' + str(nMethode) + '.txt'

def nomFichierResolutionPlottable(nomFichier, nMethode):
    """nomFichier est un string qui correspond au nom du fichier excel.
    n_methode est le numéro de la méthode.
    Renvoie le nom du fichier txt."""
    L = nomFichier.split('.')
    return 'Phase2/Solutions/Solution' + L[0][27:] + 'ByV' + str(nMethode) + 'plottable.txt'

def lignesSolution(X, h, L, TasksDico, EmployeesDico):
    """X et h sont des tableaux numpy.
    X est de dimension 3 et h de dimension 1.
    Renvoie la liste des lignes sous la forme souhaitée pour le fichier .txt."""
    Duree = vecteurDurees(TasksDico)
    premiereLigne = 'taskId;performed;employeeName;startTime;'
    listeDesLignes = [premiereLigne]
    n, _, y = X.shape
    employeeName = ''
    nombreTaches = len(TasksDico)
    for i in range(nombreTaches):
        j = 0
        tacheiAjoutee = False
        while j < y and not(tacheiAjoutee):
            numeroEmploye = 0
            while numeroEmploye < n and not(tacheiAjoutee):
                if X[numeroEmploye, i, j] == 1:
                    employeeName = EmployeesDico[numeroEmploye]['EmployeeName']
                    listeDesLignes.append(
                        'T' + str(i+1) + ';' + '1' + ';' + str(employeeName) + ';' + str(round(h[i])) + ';')
                    tacheiAjoutee = True
                numeroEmploye = numeroEmploye + 1
            j = j + 1
        if not(tacheiAjoutee):
            listeDesLignes.append(
                'T' + str(i+1) + ';' + '0' + ';' + ';' + ';')
    listeDesLignes.append(' ')  # saut de ligne
    listeDesLignes.append('employeeName;lunchBreakStartTime;')
    for numeroEmploye in range(n):
        i=0
        tachePrePauseTrouve = False
        while (i < y and not(tachePrePauseTrouve)):
            j = 0
            while (j < y and not(tachePrePauseTrouve)):
                if L[numeroEmploye, i, j] == 1:
                    listeDesLignes.append(str(
                        EmployeesDico[numeroEmploye]['EmployeeName']) + ';' + str(round(max([h[i] + Duree[i], 720]))) + ';')
                    tachePrePauseTrouve = True
                j += 1
            i += 1
    return listeDesLignes


def creationFichier(nomFichier, nMethode, X, h, L, TasksDico, EmployeesDico):
    """nomFichier est le nom du fichier utilisée pour créer les variables globales.
    n_methode est le numéro de la méthode utilisée.
    X est un tableau en 3 dimensions où chaque coefficient permet de savoir si l'employé n est allé de la tâche i à la tâche j.
    Ne renvoie rien mais crée ou modifie le fichier .txt."""
    fichier = open(nomFichierResolution(nomFichier, nMethode), "w")
    fichier.write("\n".join(lignesSolution(
        X, h, L, TasksDico, EmployeesDico)))
    fichier.close()
    return None


def performances2(tpsExec, tailleEntree, tailleMemoire, instance):
    # Ecriture des critères de performance dans un excel
    my_path = "./performance2.xlsx"
    my_wb = openpyxl.load_workbook(my_path)
    my_sheet = my_wb.active
    # on cherche à partir de quelle ligne écrire (écriture à la suite)
    i = 4
    cell = my_sheet.cell(row=i, column=1)
    while cell.value != None:
        i += 1
        cell = my_sheet.cell(row=i, column=1)
    # on ajoute les valeurs de performance obtenue du code
    cell.value = tpsExec  # temps d'execution en première colonne
    cell = my_sheet.cell(row=i, column=2)
    cell.value = tailleEntree  # taille des instances d'entrée en deuxième colonne
    cell = my_sheet.cell(row=i, column=3)
    # taille de la mémoire occupée par le programme en troisième colonne
    cell.value = tailleMemoire
    # on calcule àpartir de ces valeurs de nouveaux indicateurs
    cell = my_sheet.cell(row=i, column=4)
    cell.value = tpsExec/tailleEntree
    cell = my_sheet.cell(row=i, column=5)
    cell.value = tailleMemoire/tailleEntree
    cell = my_sheet.cell(row=i, column=6)
    cell.value = instance
    # on enregistre les données au sein de l'excel
    my_wb.save("./performance2.xlsx")

def dispostache(tasknb, TasksDico, TasksUnavailDico):
    '''retourne une liste des créneaux dispo de la tâche tasknb'''
    ouverture=recuperationHeure(TasksDico[tasknb]['OpeningTime'])
    fermeture=recuperationHeure(TasksDico[tasknb]['ClosingTime'])
    dispos=[[ouverture]]
    k=0
    for t in TasksUnavailDico:
        if t['TaskId']=='T'+str(tasknb+1):
            k+=1
            dispos[k-1].append(recuperationHeure(t['Start']))
            dispos.append([recuperationHeure(t['End'])])
    dispos[-1].append(fermeture)
    return dispos

def lignesSolutionPlottable(X, h, L, TasksDico, EmployeesDico,EmployeesUnavailDico):
    """X et h sont des tableaux numpy.
    X est de dimension 3 et h de dimension 1.
    Renvoie la liste des lignes sous la forme souhaitée pour le fichier .txt."""
    Duree = vecteurDurees(TasksDico)
    premiereLigne = 'taskId;performed;employeeName;startTime;'
    listeDesLignes = [premiereLigne]
    n, _, y = X.shape
    employeeName = ''
    nombreTaches = len(TasksDico)
    for i in range(nombreTaches):
        j = 0
        tacheiAjoutee = False
        while j < y and not(tacheiAjoutee):
            numeroEmploye = 0
            while numeroEmploye < n and not(tacheiAjoutee):
                if X[numeroEmploye, i, j] == 1:
                    employeeName = EmployeesDico[numeroEmploye]['EmployeeName']
                    listeDesLignes.append(
                        'T' + str(i+1) + ';' + '1' + ';' + str(employeeName) + ';' + str(round(h[i])) + ';')
                    tacheiAjoutee = True
                numeroEmploye = numeroEmploye + 1
            j = j + 1
        if not(tacheiAjoutee):
            listeDesLignes.append(
                'T' + str(i+1) + ';' + '0' + ';' + ';' + ';')
    for Unavail in EmployeesUnavailDico:
        ligne=('I' + '0' + ';' + '1' + ';' + Unavail['EmployeeName'] + ';' + str(recuperationHeure(Unavail['Start'])) + ';')
        listeDesLignes.append(ligne)

    listeDesLignes.append(' ')  # saut de ligne
    listeDesLignes.append('employeeName;lunchBreakStartTime;')
    for numeroEmploye in range(n):
        i=0
        tachePrePauseTrouve = False
        while (i < y and not(tachePrePauseTrouve)):
            j = 0
            while (j < y and not(tachePrePauseTrouve)):
                if L[numeroEmploye, i, j] == 1:
                    listeDesLignes.append(str(
                        EmployeesDico[numeroEmploye]['EmployeeName']) + ';' + str(round(max([h[i] + Duree[i], 720]))) + ';')
                    tachePrePauseTrouve = True
                j += 1
            i += 1
    return listeDesLignes


def creationFichierPlottable(nomFichier, nMethode, X, h, L, TasksDico, EmployeesDico, EmployeesUnavailDico):
    """nomFichier est le nom du fichier utilisée pour créer les variables globales.
    n_methode est le numéro de la méthode utilisée.
    X est un tableau en 3 dimensions où chaque coefficient permet de savoir si l'employé n est allé de la tâche i à la tâche j.
    Ne renvoie rien mais crée ou modifie le fichier .txt."""
    fichier = open(nomFichierResolutionPlottable(nomFichier, nMethode), "w")
    fichier.write("\n".join(lignesSolutionPlottable(
        X, h, L, TasksDico, EmployeesDico,EmployeesUnavailDico)))
    fichier.close()
    return None
