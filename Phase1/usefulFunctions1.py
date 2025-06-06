# Import de modules
import numpy as np
import pandas as pd
import math
import openpyxl


# Fonction récupération des données issues des excel


def extractionData(path):
    """Permet l'extraction des données depuis un fichier excel.
    path est une chaîne de caratères correspondant à l'emplacement relatif du fichier.
    Renvoie les différents dictionnaires permettant un traitement des données avec python."""
    xls = pd.ExcelFile(path)
    df1 = pd.read_excel(xls, 'Employees')
    df2 = pd.read_excel(xls, 'Tasks')

    # Création des dictionnaires de données
    employeesDico = df1.to_dict('records')
    tasksDico = df2.to_dict('records')

    return employeesDico, tasksDico


# Fonctions utiles pour créer les matrices de données utiles

def deg2rad(dd):
    """Convertit un angle "degrés décimaux" en "radians"
    """
    return dd/180*math.pi

def distanceGPS(latA, longA, latB, longB):
    """Retourne la distance en mètres entre les 2 points A et B connus grâce à
       leurs coordonnées GPS (en radians).
    """
    # Rayon de la terre en mètres (sphère IAG-GRS80)
    RT = 6378137
    # angle en radians entre les 2 points
    x = math.sin(latA)*math.sin(latB) + math.cos(latA)*math.cos(latB)*math.cos(abs(longB-longA))        
    if abs(x-1) <= 0.000000000001:
        x = 1
    elif abs(x+1) <= 0.000000000001:
        x = -1
    S = math.acos(x)
    # distance entre les 2 points, comptée sur un arc de grand cercle
    return S*RT

def distance(id1, id2, TasksDico):
    """Calcule la distance entre deux points dont on connait les coordonnées GPS.
    Entrée : les taskid correspondantes et le dictionnaire de données.
    Sortie : distance en km."""
    foundId1, foundId2 = False, False
    index = 0
    while not (foundId1 and foundId2):
        if TasksDico[index]['TaskId'] == id1:
            foundId1 = True
            long1 = deg2rad(TasksDico[index]['Longitude'])
            lat1 = deg2rad(TasksDico[index]['Latitude'])
        if TasksDico[index]['TaskId'] == id2:
            foundId2 = True
            long2 = deg2rad(TasksDico[index]['Longitude'])
            lat2 = deg2rad(TasksDico[index]['Latitude'])
        index += 1

    #distance d'arc entre deux points
    distance = round(distanceGPS(lat1,long1,lat2,long2)/1000,1)
    return distance


def competenceOK(employeeName, taskId, tasksDico, employeesDico):
    """EmployeeName est une chaîne de caractères correspondant au nom d'un employé.
    TaskId est l'identifiant d'une tâche.
    TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Retourne 1 si l'employé a la bonne compétence et un niveau suffisant pour effectuer la tâche, 0 sinon."""
    taskSkill = next(item['Skill']
                     for item in tasksDico if item['TaskId'] == taskId)
    employeeSkill = next(
        item['Skill'] for item in employeesDico if item['EmployeeName'] == employeeName)
    taskLevel = next(item['Level']
                     for item in tasksDico if item['TaskId'] == taskId)
    employeeLevel = next(
        item['Level'] for item in employeesDico if item['EmployeeName'] == employeeName)
    if taskLevel <= employeeLevel and taskSkill == employeeSkill:
        return 1
    else:
        return 0


# Fonctions création des matrices nécessaires au calcul de la fonction coût et à la vérification des contraintes

def matriceDistance(tasksDico):
    """TasksDico est le dictionnaire contenant les informations sur les tâches.
    Renvoie la matrice des distances entre les différentes tâches."""
    matriceDesDistances = []
    for tache1 in tasksDico:
        taskId1 = tache1["TaskId"]
        ligne = []
        for tache2 in tasksDico:
            taskId2 = tache2["TaskId"]
            ligne.append(distance(taskId1, taskId2, tasksDico))
        matriceDesDistances.append(ligne)
    return matriceDesDistances

d = matriceDistance(extractionData("Phase1/InstancesV1/InstancePolandV1.xlsx")[1])
# print(d)
print(d[6][14])

def matriceCompetences(employeesDico, tasksDico):
    """TasksDico et EmployeesDico sont les dictionnaires contenant les informations respectives sur les tâches et les employés.
    Renvoie une matrice C qui comporte un 1 si l'employé n peut faire la tache i et un 0 sinon."""
    nombreEmployes = len(employeesDico)
    nombreTaches = len(tasksDico)
    C = np.zeros((nombreEmployes, nombreTaches))
    for n in range(nombreEmployes):
        for i in range(nombreTaches):
            C[n, i] = competenceOK(
                employeesDico[n]['EmployeeName'], tasksDico[i]['TaskId'], tasksDico, employeesDico)
    return C


# Fonctions récupération des contraintes temporelles des tâches

def recuperationHeure(heure):
    """heure est une chaîne de caractères de la forme 11:15pm.
    Renvoie un entier correspondant au nombre de minutes écoulées depuis minuit."""
    res = heure.split(':')
    h = int(res[0])  # l'heure
    m = int(res[1][:2])  # les minutes
    if res[1][2:] == 'pm':
        if res[0] != 12:
            h += 12  # modifications pour l'aprem
    else:
        if res[0] == 12:
            h = 0
    return h*60+m


def vecteurOuvertures(tasksDico):
    """TasksDico est le dictionnaire contenant les informations respectives sur les tâches.
    Renvoie une liste avec l'heure d'ouverture de chaque tâche."""
    nombreTaches = len(tasksDico)
    O = []
    for i in range(nombreTaches):
        heureEnMinutes = recuperationHeure(tasksDico[i]['OpeningTime'])
        O.append(heureEnMinutes)
    return O


def vecteurFermetures(tasksDico):
    """TasksDico est le dictionnaire contenant les informations respectives sur les tâches.
    Renvoie une liste avec l'heure de fermeture de chaque tâche."""
    nombreTaches = len(tasksDico)
    F = []
    for i in range(nombreTaches):
        heureEnMinutes = recuperationHeure(tasksDico[i]['ClosingTime'])
        F.append(heureEnMinutes)
    return F


def vecteurDurees(tasksDico):
    """TasksDico est le dictionnaire contenant les informations sur les tâches.
    Renvoie le vecteur des durées de chaque tâche."""
    vecteurDuree = []
    for row in tasksDico:
        vecteurDuree.append(row['TaskDuration'])
    return vecteurDuree


# Création du fichier solution


def nomFichierResolution(nomFichier, nMethode):
    """nom_fichier est un string qui correspond au nom du fichier excel.
    n_methode est le numéro de la méthode.
    Renvoie le nom du fichier txt."""
    L = nomFichier.split('.')
    nMethode = str(nMethode)
    # return 'Solution' + L[0][8:] + 'ByV' + n_methode + '.txt' # ancienne version sans arborescence pour les fichiers excels
    # nouvelle version avec arborescence pour les fichiers excels
    return 'Phase1/Solutions/Solution' + L[0][27:] + 'ByV' + nMethode + '.txt'


def lignesSolution(X, h, employeesDico):
    """X et h sont des tableaux numpy.
    X est de dimension 3 et h de dimension 1.
    Renvoie la liste des lignes sous la forme souhaitée pour le fichier .txt."""
    premiereLigne = 'taskId;performed;employeeName;startTime;'
    listeDesLignes = [premiereLigne]
    n, x, y = X.shape
    employeeName = ''
    for i in range(x - 2*n):
        j = 0
        tacheiAjoutee = False
        while j < y and not(tacheiAjoutee):
            numeroEmploye = 0
            while numeroEmploye < n and not(tacheiAjoutee):
                if X[numeroEmploye, i, j] == 1:
                    employeeName = employeesDico[numeroEmploye]['EmployeeName']
                    listeDesLignes.append(
                        'T' + str(i+1) + ';' + '1' + ';' + str(employeeName) + ';' + str(round(h[i])) + ';')
                    tacheiAjoutee = True
                numeroEmploye = numeroEmploye + 1
            j = j + 1
        if not(tacheiAjoutee):
            listeDesLignes.append(
                'T' + str(i+1) + ';' + '0' + ';' + ';' + ';')
    return listeDesLignes


def creationFichier(nomFichier, nMethode, X, h, employeesDico):
    """nomFichier est le nom du fichier utilisée pour créer les variables globales.
    n_methode est le numéro de la méthode utilisée.
    X est un tableau en 3 dimensions où chaque coefficient permet de savoir si l'employé n est allé de la tâche i à la tâche j.
    Ne renvoie rien mais crée ou modifie le fichier .txt."""
    fichier = open(nomFichierResolution(nomFichier, nMethode), "w")
    fichier.write("\n".join(lignesSolution(X, h, employeesDico)))
    fichier.close()
    return None


def performances1(tpsExec, tailleEntree, tailleMemoire, instance):
    # Ecriture des critères de performance dans un excel
    myPath = "./performance1.xlsx"
    myWb = openpyxl.load_workbook(myPath)
    mySheet = myWb.active
    # on cherche à partir de quelle ligne écrire (écriture à la suite)
    i = 4
    cell = mySheet.cell(row=i, column=1)
    while cell.value != None:
        i += 1
        cell = mySheet.cell(row=i, column=1)
    # on ajoute les valeurs de performance obtenue du code
    cell.value = tpsExec  # temps d'execution en première colonne
    cell = mySheet.cell(row=i, column=2)
    cell.value = tailleEntree  # taille des instances d'entrée en deuxième colonne
    cell = mySheet.cell(row=i, column=3)
    # taille de la mémoire occupée par le programme en troisième colonne
    cell.value = tailleMemoire
    # on calcule àpartir de ces valeurs de nouveaux indicateurs
    cell = mySheet.cell(row=i, column=4)
    cell.value = tpsExec/tailleEntree
    cell = mySheet.cell(row=i, column=5)
    cell.value = tailleMemoire/tailleEntree
    cell = mySheet.cell(row=i, column=6)
    cell.value = instance
    # on enregistre les données au sein de l'excel
    myWb.save("./performance1.xlsx")
